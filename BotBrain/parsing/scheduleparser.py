import openpyxl
from loguru import logger

from .. import sql as sqlapi
from ..utils.text import clear_invisible_character, check_for_group_tag

ROW_WITH_GROUP_NAMES = 3
COL_WITH_GROUP_NAMES = 3
MAX_ROW = ROW_WITH_GROUP_NAMES + 72  # 12 строк - 1 день. 16*6 дней = 72
LEFORTOVO_COLORS = ['FFE1FFE1', 'FFCCFFCC']


def __find_groups_and_cabs_columns(sheet):
    all_groups_cabs = []  # состоит из таких подмассивов [('ТА-11', 'group_col'),'cab_col']
    last_group = ''
    for groups_and_cabs in sheet.iter_rows(min_row=ROW_WITH_GROUP_NAMES, max_row=ROW_WITH_GROUP_NAMES,
                                           min_col=COL_WITH_GROUP_NAMES, max_col=sheet.max_column,
                                           values_only=True):
        current_col = COL_WITH_GROUP_NAMES - 1  # минус один потому что в цикле сразу идет увеличение на единицу
        for gc in groups_and_cabs:
            current_col += 1
            if check_for_group_tag(gc):
                if gc == last_group:
                    continue
                all_groups_cabs.append([(gc, current_col)])
                last_group = gc
            elif gc in ('каб.', 'каб'):
                all_groups_cabs[-1].append(current_col)

    return all_groups_cabs


# берет столбец с переданный кабинетами и возвращает все кабинеты на один день с разделением на недели (верх/низ)
def __get_cabs(sheet, day, column):
    upper_cabs = []
    lower_cabs = []
    """
    min_row - начальная строка, начинается со строки под названием групп (ROW_WITH_GROUP_NAMES + 1) + 12 строк за
        каждый день (12 * (day - 1)), поскольку в первый день сдвигать не нужно (- 1)
    max_row - максимальная строка на 12 больше начальной
    """
    for cabinets in sheet.iter_rows(min_row=ROW_WITH_GROUP_NAMES + 1 + 12 * (day - 1),
                                    max_row=ROW_WITH_GROUP_NAMES + 12 * day,
                                    min_col=column, max_col=column,
                                    values_only=False):
        for cab in cabinets:
            if True if cab.row % 2 == 0 else False:
                upper_cabs.append(cab.value if cab.value is not None else '')
            else:
                lower_cabs.append(cab.value if cab.value is not None else '')
    return upper_cabs, lower_cabs


def __get_lessons(sheet):
    all_groups_cabs = __find_groups_and_cabs_columns(sheet)
    # all_lessons [group_name, day, isupper, islef, (lessons_mas, cab_mas)] 6 раз (6 дней) для каждой группы
    all_lessons = []
    for gc in all_groups_cabs:
        group_name = gc[0][0]
        day = 1
        islef = False

        upper_lessons = []
        lower_lessons = []
        # проход по строчкам в столбце с группой с линии под названием группы и до конца
        for lessons in sheet.iter_rows(min_row=ROW_WITH_GROUP_NAMES + 1, max_row=MAX_ROW,
                                       min_col=gc[0][1], max_col=gc[0][1],
                                       values_only=False):
            for lesson_name in lessons:
                # верхняя/нижняя неделя
                isupper = True if lesson_name.row % 2 == 0 else False
                # определение лефотово или нет по цвету клетки
                islef = True if lesson_name.fill.start_color.rgb in LEFORTOVO_COLORS else islef

                if isupper:
                    upper_lessons.append(clear_invisible_character(lesson_name.value, separator=' '))
                else:
                    lower_lessons.append(clear_invisible_character(lesson_name.value, separator=' '))

                # Если отбросить первые три строчки (-3), то каждые 12 строк меняется день
                if (lesson_name.row - 3) % 12 == 0:
                    up_cabs, low_cabs = __get_cabs(sheet, day, gc[1])

                    upper_lessons = list(zip(upper_lessons, up_cabs))
                    lower_lessons = list(zip(lower_lessons, low_cabs))

                    #                  [group_name, day, isupper, islef, lessons_mas]
                    all_lessons.append([group_name, day, True, islef, upper_lessons])
                    all_lessons.append([group_name, day, False, islef, lower_lessons])

                    upper_lessons = []
                    lower_lessons = []

                    islef = False
                    day += 1
    return all_lessons


def update_schedule(path_to_file):
    wb = openpyxl.load_workbook(filename=path_to_file, read_only=True)
    sheet = wb[wb.sheetnames[0]]

    lessons_data = __get_lessons(sheet)

    with sqlapi.mysqlapiwrapper() as db:
        db.schedule.clear()
        for data in lessons_data:
            db.schedule.insert(group=data[0], day=data[1], isupper=int(data[2]),
                               islef=int(data[3]), lessons=data[4])
    wb.close()
    logger.info('Schedule updated')
